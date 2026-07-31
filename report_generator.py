# report_generator.py
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

def generate_excel_report(df, output_path):
    wb = openpyxl.Workbook()

    # --- TAB 1: EXECUTIVE SUMMARY ---
    ws_summary = wb.active
    ws_summary.title = "Executive Summary"
    
    total_records = len(df)
    total_anomalies = int(df["is_anomaly"].sum())
    total_overbilling = float(df[df["potential_financial_leakage"] > 0]["potential_financial_leakage"].sum())
    total_unbilled = float(abs(df[df["potential_financial_leakage"] < 0]["potential_financial_leakage"].sum()))
    net_exposure = total_overbilling - total_unbilled

    ws_summary.append(["ENTERPRISE REVENUE LEAKAGE AUDIT REPORT"])
    ws_summary.append([f"Execution Timestamp: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
    ws_summary.append([])
    
    ws_summary.append(["KPI Metric", "Value"])
    summary_metrics = [
        ["Total Shifts Audited", total_records],
        ["Total Anomalies Identified", total_anomalies],
        ["Total Potential Overbilled Leakage ($)", round(total_overbilling, 2)],
        ["Total Unbilled Work Exposure ($)", round(total_unbilled, 2)],
        ["Net Financial Exposure ($)", round(net_exposure, 2)]
    ]
    
    for row in summary_metrics:
        ws_summary.append(row)

    # Executive Formatting
    ws_summary["A1"].font = Font(size=15, bold=True, color="1F497D")
    ws_summary["A2"].font = Font(size=9, italic=True, color="595959")

    for col in range(1, 3):
        cell = ws_summary.cell(row=4, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # --- TAB 2: AUDIT DETAILS ---
    ws_details = wb.create_sheet(title="Audit Details Log")
    
    for r in dataframe_to_rows(df, index=False, header=True):
        ws_details.append(r)

    # Detail Headers
    header_fill = PatternFill(start_color="244062", end_color="244062", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for cell in ws_details[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Conditional Highlighting for Flagged Rows
    flag_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    flag_font = Font(color="C00000", bold=True)
    audit_flag_col = df.columns.get_loc("audit_flag") + 1

    for row in range(2, ws_details.max_row + 1):
        cell = ws_details.cell(row=row, column=audit_flag_col)
        if cell.value != "CLEAN":
            cell.fill = flag_fill
            cell.font = flag_font

    # Column Auto-Widths
    for ws in [ws_summary, ws_details]:
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    wb.save(output_path)